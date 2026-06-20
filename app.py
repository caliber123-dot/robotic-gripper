# from turtle import mode
import math
from io import BytesIO
import os
from flask import Flask, render_template, request, jsonify, send_file
from waitress import serve as waitress_serve
from datetime import datetime
import time
from merge_excel import fileMerged

# from openpyxl import Workbook
# from openpyxl.worksheet.worksheet import Worksheet
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from sympy import symbols, diff, sympify, N
from sympy.core.expr import Expr
from shape_drawer import (
    generate_sphere,
    generate_sphere_3d_2,
    generate_rectangle,
    generate_rectangle_3d,
    generate_ellipsoid,
    generate_ellipsoid_3d_2,
    generate_rectangle_vertical,
    generate_rectangle_vertical_3d,
    generate_ellipsoid_vertical,
    generate_ellipsoid_vertical_3d,
)

import base64
from io import BytesIO
from openpyxl.drawing.image import Image

# from openpyxl.styles

# from reportlab.pdfgen import canvas
from reportlab.platypus import (
    SimpleDocTemplate,
    Paragraph,
    Spacer,
    Table,
    TableStyle,
    Image as RLImage,
)
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.pagesizes import letter, landscape

from reportlab.pdfbase import pdfmetrics
from reportlab.pdfbase.ttfonts import TTFont

# pdfmetrics.registerFont(TTFont("Arial", "arial.ttf"))
# pdfmetrics.registerFont(TTFont("Arial-Bold", "arialbd.ttf"))
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
pdfmetrics.registerFont(
    TTFont("Arial-Bold", os.path.join(BASE_DIR, "static/css/fonts/arialbd.ttf"))
)
from database import (
    create_tables,
    save_input,
    get_saved_input,
    get_saved_input_graph,
    is_duplicate,
    update_input,
    save_spring_constants,
    get_spring_constants,
    get_saved_input_graph_all,
    get_comparison_time,
    get_comparison_time1,
    get_comparison_time2,
    get_spring_constants_comparison,
    get_spring_constants_comparison2,
    get_saved_input_compare,
    get_saved_input_compare2,
    get_comparison_all_equal,
)

app = Flask(__name__)

UPLOAD_FOLDER = os.path.join(app.root_path, "static", "files")
os.makedirs(UPLOAD_FOLDER, exist_ok=True)

create_tables()


@app.route("/")
def hello_world():
    return render_template("landing.html")


# Home Screen
@app.route("/home")
def home():
    # print("############# Rendering index.html")
    return render_template("home.html")


@app.route("/input")
def input():
    # print("############# Rendering index.html")
    return render_template("index.html")


@app.route("/graph")
def graph():
    # print("############# Rendering Bar Chart graph.html")
    return render_template("graph.html")


@app.route("/comparison")
def comparison():
    # print("############# Rendering Bar Chart graph.html")
    return render_template("comparison.html")


# Material density
materials = {"rubber": 1100, "abs": 1040, "teflon": 2200}


# Volume calculation
def calculate_volume(shape):
    if shape == 1:  # Rectangular
        return 0.1 * 0.04 * 0.02
    elif shape == 2:  # Sphere
        return (4 / 3) * 3.14 * (0.05**3)
    elif shape == 3:  # Ellipsoid
        return (4 / 3) * 3.14 * 0.05 * 0.03 * 0.03
    else:
        return 0  # 🔥 VERY IMPORTANT


def parse_theta_function(func_str, t_value):

    # normalize input
    func_str = func_str.replace("^", "**").replace(" ", "")

    t = symbols("t")

    # symbolic expression
    expr: Expr = sympify(func_str)

    # derivatives
    d1: Expr = diff(expr, t)
    d2: Expr = diff(d1, t)

    # numerical evaluation
    th = float(N(expr.subs(t, t_value)))
    dth = float(N(d1.subs(t, t_value)))
    d2th = float(N(d2.subs(t, t_value)))

    return th, dth, d2th


# θ(t) and c = func
def theta(c, t):
    if c == 1:
        return t / 2 + (t * t) / 3
    elif c == 2:
        return t / 2 + (t * t) / 3 + (t * t * t) / 4
    elif c == 3:
        return t / 3 + (t * t) / 4 + (t * t * t) / 5


# θ'(t)
def dtheta(c, t):
    if c == 1:
        return 1 / 2 + (2 * t) / 3
    elif c == 2:
        return 1 / 2 + (2 * t) / 3 + (3 * t * t) / 4
    elif c == 3:
        return 1 / 3 + (2 * t) / 4 + (3 * t * t) / 5


# θ''(t)
def d2theta(c, t):
    if c == 1:
        return 2 / 3
    elif c == 2:
        return 2 / 3 + (3 * t) / 2
    elif c == 3:
        return 1 / 2 + (6 * t) / 5


# ===================== EQ 14 =====================
g = 9.8  # constant from your notes in Part A
r = 0.05  # constant from your notes in Part A and B


def compute_A(M, th, dth, d2th):
    # mgr2 = d2th  - th * d2th- (th**2 / 2) * dth + (th**3 / 6) * dth + (th**4 / 24) * d2th - (th**5 / 120) * dth
    # print("mgr2:", mgr2)
    return (M * g * r) * (
        d2th
        - th * dth
        - (th**2 / 2) * d2th
        + (th**3 / 6) * dth
        + (th**4 / 24) * d2th
        - (th**5 / 120) * dth
    )


def compute_B(th):
    return r * (th - (th**3) / 6 + (th**5) / 120)


# F = A + K * B
def cal_force_eq14(M, K, func, t):
    # CALL DYNAMIC FUNCTION PARSER
    # θ(t), θ′(t), θ″(t)
    th, dth, d2th = parse_theta_function(func, t)

    A = compute_A(M, th, dth, d2th)
    B = compute_B(th)

    return A, B, K  # ✅ RETURN SEPARATE VALUES


# ===================== New A1, A2 and B1, B2==========================
def compute_A1A2(M, th, dth, d2th):
    # A1 = (mg)i
    A1 = M * g
    # A2 equation
    # A2 = (
    #     - r * math.sin(th) * dth
    #     + r * math.cos(th) * d2th
    # )
    A2 = r * (
        d2th
        - th * dth
        - ((th**2) / 2) * d2th
        + ((th**3) / 6) * dth
        + ((th**4) / 24) * d2th
        - ((th**5) / 120) * dth
    )
    # print("A1:", A1)
    # print("Computed A2:", A2)
    # Print - r * math.sin(th) * dth
    #     + r * math.cos(th) * d2th
    # print("r:", r, "th:", th, "dth:", dth, "d2th:", d2th, "sin(th):", math.sin(th), "cos(th):", math.cos(th))

    return A1, A2


def compute_B1B2(th):
    # B1 = ri
    B1 = r
    # B2 Taylor series
    B2 = th - (th**3) / 6 + (th**5) / 120
    return B1, B2


def cal_force_eq14A1B1(M, K, func, t):
    # CALL DYNAMIC FUNCTION PARSER
    # θ(t), θ′(t), θ″(t)
    th, dth, d2th = parse_theta_function(func, t)

    A1, A2 = compute_A1A2(M, th, dth, d2th)

    B1, B2 = compute_B1B2(th)

    A = A1 * A2
    B = B1 * B2

    return A1, A2, B1, B2, A, B, K


@app.route("/calculate", methods=["POST"])
def calculate():
    start_time = time.time()  # ⏱ start
    # print(request.json)
    data = request.json
    # shape = int(data["shape"])
    shape = int(data.get("shape", 0))
    event = data["event"]  # 🔥 GET EVENT TYPE string value

    length = float(data.get("length", 0)) * 0.001
    breadth = float(data.get("breadth", 0)) * 0.001
    width = float(data.get("width", 0)) * 0.001

    radius = float(data.get("radius", 0)) * 0.001

    Rmajor = float(data.get("Rmajor", 0)) * 0.001
    Rminor = float(data.get("Rminor", 0)) * 0.001

    material = data["material"]
    t = float(data["time"])
    func = data["func"]
    # print("Function:", func)
    # print("Selected Function ID:", func)
    gripper = int(data["gripper"])  # 1 or 2
    # print("Gripper type:", "4-Finger Gripper" if gripper == 1 else "3-Finger Gripper with Thumb")

    if shape == 1:  # Rectangular
        if event == "length":
            # L, B, H = 0.1, 0.04, 0.02
            L, B, H = length, breadth, width
        elif event == "breadth":
            # L, B, H = 0.04, 0.1, 0.02
            L, B, H = breadth, length, width
        else:
            L, B, H = 0.1, 0.04, 0.02  # default ✅
        volume = L * B * H
        # print("Volume:", volume)

    elif shape == 2:  # Spherical
        # r = 0.05
        r = radius
        volume = (4 / 3) * math.pi * (r**3)

    elif shape == 3:  # Ellipsoid
        if event == "major":
            # a, b, c = 0.05, 0.03, 0.03
            a, b, c = Rmajor, Rminor, Rminor
        elif event == "minor":
            # a, b, c = 0.03, 0.03, 0.05
            a, b, c = Rminor, Rminor, Rmajor
        else:
            a, b, c = 0.05, 0.03, 0.03  # default ✅
        volume = (4 / 3) * math.pi * a * b * c

    else:
        volume = 0
    density = materials.get(material, 0)
    M = density * volume
    # print("Density:", density)
    # print("Volume:", volume)
    # print("Mass:", M)

    forces = []
    # forcesA = []
    # forcesB = []
    forcesA1 = []
    forcesA2 = []
    forcesB1 = []
    forcesB2 = []
    kf_total = []
    thumb = 0
    ktt = 0
    tA1 = 0
    tA2 = 0
    tB1 = 0
    tB2 = 0
    finger_count = 4 if gripper == 1 else 3
    mode = data["mode"]  # get string constant values eg 1, 2, 3
    spring_data = []
    # print("\n===== CALCULATION MODE =====")
    # ================= MODE 1 All equal =================
    if mode == "1":  # All equal
        k = float(data.get("k_common", 0))
        spring_data.append({"spring_key": "k_common", "spring_value": k})
        kf = (k * k) / (k + k)

        # Fingers
        for i in range(finger_count):
            # 🔥 forces.append(round(calculate_force(M, k_finger_total, func, t), 3))
            kf_total.append(kf)
            # A, B, K = cal_force_eq14(M, kf, func, t)
            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, kf, func, t)
            # F = A + K * B
            F = abs(A) + K * abs(B)  # ✅ Use absolute values to avoid negative forces
            forcesA1.append(round(abs(A1), 4))
            forcesA2.append(round(abs(A2), 4))
            forcesB1.append(round(abs(B1), 4))
            forcesB2.append(round(abs(B2), 4))
            forces.append(round(F, 4))

        # Thumb (only if exists)
        if gripper == 2:
            # thumb = round(calculate_force(M, k_thumb_total, func, t), 3)
            ktt = 1 / ((1 / k) + (1 / k) + (1 / k))  # Parallel spring formula
            # A, B, K = cal_force_eq14(M, ktt, func, t)
            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, ktt, func, t)
            F = abs(A) + K * abs(B)
            tA1 = round(abs(A1), 4)
            tA2 = round(abs(A2), 4)
            tB1 = round(abs(B1), 4)
            tB2 = round(abs(B2), 4)
            # print("Thumb A:", tA)
            # print("Thumb B:", tB)
            # print("Thumb K:", ktt)
            thumb = round(F, 4)

    # ================= MODE 2 Fingers same, Thumb different=================
    elif mode == "2":
        kf = float(data.get("k_finger", 0))
        spring_data.append({"spring_key": "k_finger", "spring_value": kf})
        # 🔥 FIX: Finger has 2 springs → multiply by 2
        # kf_total = kf * 2
        kf = (kf * kf) / (kf + kf)
        # ✅ Store total finger stiffness for debugging
        # print("kf_total:", kf_total)
        for i in range(finger_count):
            # forces.append(round(calculate_force(M, kf_total, func, t), 2))
            # cal_force_eq14(M, K, func, t)
            kf_total.append(kf)
            # A, B, K = cal_force_eq14(M, kf, func, t)
            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, kf, func, t)
            F = abs(A) + K * abs(B)
            forcesA1.append(round(abs(A1), 4))
            forcesA2.append(round(abs(A2), 4))
            forcesB1.append(round(abs(B1), 4))
            forcesB2.append(round(abs(B2), 4))
            forces.append(round(F, 4))

            # forces.append(round(cal_force_eq14(M, kf_total, func, t), 4))
        if gripper == 2:
            # kt = sum(data.get("thumb", []))  # already correct (108+112+100=320)
            kt = data.get("thumb", [])
            # spring_data.extend(kt)
            if len(kt) > 0:
                spring_data.append({"spring_key": "k_thumb", "spring_value": kt[0]})

            if len(kt) > 1:
                spring_data.append({"spring_key": "k_thumb2", "spring_value": kt[1]})

            if len(kt) > 2:
                spring_data.append({"spring_key": "k_thumb3", "spring_value": kt[2]})
            if len(kt) >= 3:
                ktt = 1 / (
                    (1 / kt[0]) + (1 / kt[1]) + (1 / kt[2])
                )  # Parallel spring formula
                # print("ktt:", ktt)
            # A, B, K = cal_force_eq14(M, ktt, func, t)
            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, ktt, func, t)
            F = abs(A) + K * abs(B)
            tA1 = round(abs(A1), 4)
            tA2 = round(abs(A2), 4)
            tB1 = round(abs(B1), 4)
            tB2 = round(abs(B2), 4)
            # print("Thumb A:", tA)
            # print("Thumb B:", tB)
            # print("Thumb K:", ktt)
            thumb = round(F, 4)
            # thumb = round(cal_force_eq14(M, ktt, func, t), 4)

    # ================= MODE 3 All unequal=================
    elif mode == "3":
        # kfinger = []
        for i in range(finger_count):
            # k = data["fingers"][i]["k1"] + data["fingers"][i]["k2"]
            k1 = data["fingers"][i]["k1"]
            k2 = data["fingers"][i]["k2"]
            spring_data.append({"spring_key": f"f{i+1}k1", "spring_value": k1})
            spring_data.append({"spring_key": f"f{i+1}k2", "spring_value": k2})
            kf = (k1 * k2) / (k1 + k2)
            kf_total.append(kf)
            # forces.append(round(calculate_force(M, k, func, t), 2))
            # A, B, K = cal_force_eq14(M, kf, func, t)
            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, kf, func, t)
            F = abs(A) + K * abs(B)
            forcesA1.append(round(abs(A1), 4))
            forcesA2.append(round(abs(A2), 4))
            forcesB1.append(round(abs(B1), 4))
            forcesB2.append(round(abs(B2), 4))
            forces.append(round(F, 4))

        if gripper == 2:
            # kt = sum(data.get("thumb", []))
            kt = data.get("thumb", [])
            # spring_data.extend(kt)
            if len(kt) > 0:
                spring_data.append({"spring_key": "k_thumb", "spring_value": kt[0]})

            if len(kt) > 1:
                spring_data.append({"spring_key": "k_thumb2", "spring_value": kt[1]})

            if len(kt) > 2:
                spring_data.append({"spring_key": "k_thumb3", "spring_value": kt[2]})
            if len(kt) >= 3:
                ktt = 1 / (
                    (1 / kt[0]) + (1 / kt[1]) + (1 / kt[2])
                )  # Parallel spring formula
            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, ktt, func, t)
            F = abs(A) + K * abs(B)
            tA1 = round(abs(A1), 4)
            tA2 = round(abs(A2), 4)
            tB1 = round(abs(B1), 4)
            tB2 = round(abs(B2), 4)
            # print("Thumb A:", tA)
            # print("Thumb B:", tB)
            # print("Thumb K:", ktt)
            thumb = round(F, 4)
            # thumb = round(calculate_force(M, kt, func, t), 2)

    # TOTAL
    total = round(sum(forces) + thumb, 4)

    # total = round(sum(forces) + thumb, 2)
    fig1 = ""  # 🔥 IMPORTANT: Initialize variables to avoid reference errors
    # fig3d = ""
    fig3 = ""
    gripper_name = ""
    # shape_name = ""
    if gripper == 1:
        # print("gripper type: 4-Finger Gripper")
        fig1 = "static/img/model_4Fingers.png"
        fig3 = "static/img/Industrial_For4inger.jpg"
        gripper_name = "4-Finger Gripper"
    elif gripper == 2:
        # print("gripper type: 3-Finger Gripper with Thumb")
        fig1 = "static/img/model_3Fingers_Thumb.png"
        fig3 = "static/img/basic.avif"
        gripper_name = "3-Finger Gripper with Thumb"

    # ==========================================
    # SAVE SPRING VALUES TO DATABASE
    # ==========================================
    # spring_values = list(set(spring_values)) # Remove duplicates

    # save_spring_constants(
    #     gripper,
    #     spring_values
    # )
    # save_spring_constants(
    #     gripper,
    #     spring_data
    # )
    save_spring_constants(gripper, shape, material, t, func, spring_data)

    # =========================
    # SAVE INPUT TO SQLITE
    # =========================

    save_data = {
        "gripper": gripper,
        "shape": shape,
        "event": event,
        "material": material,
        "time": t,
        "func": func,
        "mode": mode,
        "length": length,
        "breadth": breadth,
        "width": width,
        "radius": radius,
        "Rmajor": Rmajor,
        "Rminor": Rminor,
        "k_common": data.get("k_common"),
        "k_finger": data.get("k_finger"),
        # "k_thumb1": data.get("k_thumb"),
        # "k_thumb2": data.get("k_thumb2"),
        # "k_thumb3": data.get("k_thumb3"),
        "f1k1": data.get("fingers", [{}])[0].get("k1"),
        "f1k2": data.get("fingers", [{}])[0].get("k2"),
        "f2k1": data.get("fingers", [{}, {}])[1].get("k1"),
        "f2k2": data.get("fingers", [{}, {}])[1].get("k2"),
        "f3k1": data.get("fingers", [{}, {}, {}])[2].get("k1"),
        "f3k2": data.get("fingers", [{}, {}, {}])[2].get("k2"),
        "f4k1": data.get("fingers", [{}, {}, {}, {}])[3].get("k1"),
        "f4k2": data.get("fingers", [{}, {}, {}, {}])[3].get("k2"),
        "Thk1": data.get("thumb", [None, None, None])[0],
        "Thk2": data.get("thumb", [None, None, None])[1],
        "Thk3": data.get("thumb", [None, None, None])[2],
        "total": total,
    }
    # =========================
    # SAVE ONLY IF NEW
    # =========================

    duplicate_id = is_duplicate(save_data)

    if not duplicate_id:

        save_input(save_data)

        # print("New entry GripperID:", gripper, "GrpperName", gripper_name, "Shape:", shape_name, "Material:", material, "Time:", t, "Function:", func, "Mode:", mode)
        # with timestamp
        # print("New entry saved at:", datetime.now())

    else:
        update_input(duplicate_id, save_data)
        # print("Duplicate found → Updated ID:", duplicate_id, "GripperID:", gripper, "GrpperName", gripper_name, "Shape:", shape_name, "Material:", material, "Time:", t, "Function:", func, "Mode:", mode)
        # print("Duplicate/Updated entry at:", datetime.now())

    # Print forces for debugging
    # print(f"t={t} sec, Forces: {forces}, Thumb: {thumb}, Total: {total}")
    end_time = time.time()  # ⏱ end
    # execution_time = (end_time - start_time) * 1000  # convert to ms
    execution_time = (end_time - start_time) * 1000000  # microseconds (µs)
    # print("fig3d:", fig3d)
    return jsonify(
        {
            "volume": volume,
            "mass": M,
            "forces": forces,
            "thumb": thumb,
            "total": total,
            "execution_time": round(execution_time, 2),
            "fig1": fig1,
            # "fig2": fig2,
            "gripper_name": gripper_name,
            # "shape_name": shape_name,
            "forcesA1": forcesA1,
            "forcesA2": forcesA2,
            "forcesB1": forcesB1,
            "forcesB2": forcesB2,
            "thumbA1": tA1,
            "thumbA2": tA2,
            "thumbB1": tB1,
            "thumbB2": tB2,
            "kf_total": kf_total,
            "ktt": ktt,
            "fig3": fig3,
            # "fig3d": fig3d
        }
    )


@app.route("/GetShapes", methods=["POST"])
def GetShapes():

    data = request.json
    shape = int(data.get("shape", 0))
    event = data["event"]  # 🔥 GET EVENT TYPE string value
    # print("Event:", event)

    length = float(data.get("length", 0)) * 0.001
    breadth = float(data.get("breadth", 0)) * 0.001
    # width = float(data.get("width", 0)) * 0.001

    radius = float(data.get("radius", 0)) * 0.001

    Rmajor = float(data.get("Rmajor", 0)) * 0.001
    Rminor = float(data.get("Rminor", 0)) * 0.001

    # =====================================
    # DYNAMIC ENGINEERING FIGURES
    # =====================================
    shape_name = ""
    fig2 = ""
    fig3d = ""
    if shape == 1:
        shape_name = "Rectangular"
        if event == "length":
            fig2 = generate_rectangle(length * 1000, breadth * 1000)
            fig3d = generate_rectangle_3d(length * 1000, breadth * 1000)
        elif event == "breadth":
            fig2 = generate_rectangle_vertical(length * 1000, breadth * 1000)
            fig3d = generate_rectangle_vertical_3d(length * 1000, breadth * 1000)

    elif shape == 2:
        shape_name = "Spherical"
        fig2 = generate_sphere(radius * 1000)
        fig3d = generate_sphere_3d_2(radius * 1000)

    elif shape == 3:
        shape_name = "Ellipsoidal"
        if event == "major":
            fig2 = generate_ellipsoid(Rmajor * 1000, Rminor * 1000)
            fig3d = generate_ellipsoid_3d_2(Rmajor * 1000, Rminor * 1000)
        elif event == "minor":
            fig2 = generate_ellipsoid_vertical(Rmajor * 1000, Rminor * 1000)
            fig3d = generate_ellipsoid_vertical_3d(Rmajor * 1000, Rminor * 1000)

    return jsonify({"shape_name": shape_name, "fig2": fig2, "fig3d": fig3d})


@app.route("/get_saved_data", methods=["POST"])
def get_saved_data():

    data = request.json

    result = get_saved_input(
        data.get("gripper"),
        int(data.get("shape", 0)),
        data.get("material"),
        float(data.get("time", 0)),
        data.get("func"),
        data.get("mode"),
    )

    if result:

        return jsonify({"status": "found", "data": result})

    return jsonify({"status": "not_found"})


@app.route("/get_saved_data_graph", methods=["POST"])
def get_saved_data_graph():

    data = request.json

    result = get_saved_input_graph(
        int(data.get("gripper", 0)),
        int(data.get("shape", 0)),
        data.get("material"),
        data.get("func"),
        data.get("mode"),
    )

    if result:

        return jsonify({"status": "found", "data": result})

    return jsonify({"status": "not_found"})


# for tab 1
@app.route("/get_saved_data_compare", methods=["POST"])
def get_saved_data_compare():

    data = request.json

    result = get_saved_input_compare(
        int(data.get("gripper", 0)),
        int(data.get("shape", 0)),
        data.get("func"),
        float(data.get("time", 0)),
        data.get("mode"),
    )

    if result:

        return jsonify({"status": "found", "data": result})

    return jsonify({"status": "not_found"})


# for tab 3
@app.route("/get_saved_data_compare2", methods=["POST"])
def get_saved_data_compare2():

    data = request.json

    result = get_saved_input_compare2(
        int(data.get("gripper", 0)),
        data.get("material", 0),
        data.get("func"),
        float(data.get("time", 0)),
        data.get("mode"),
    )

    if result:

        return jsonify({"status": "found", "data": result})

    return jsonify({"status": "not_found"})


@app.route("/download_excel22", methods=["POST"])
def download_excel22():

    data = request.json

    wb = Workbook()

    ws = wb.active

    if ws is None:
        raise Exception("Worksheet not created")

    ws.title = "Gripper Results"

    # =========================
    # HEADER DETAILS
    # =========================

    header_font = Font(bold=True)
    Grippertype = data.get("gripper_name", "")
    ws.merge_cells("A1:I1")
    ws["A1"] = "Schematic Representation of " f"Spring Structure Model ({Grippertype})"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center")
    # ws.column_dimensions["A"].width = 15
    ws["A3"] = "Gripper type:"
    ws["A3"].font = header_font
    ws["B3"] = Grippertype

    ws["A4"] = "Object Shape"
    ws["A4"].font = header_font
    ws["B4"] = data.get("shape_name", "")

    ws["A5"] = "Material"
    ws["A5"].font = header_font
    ws["B5"] = data.get("material", "")

    ws["A6"] = "Time"
    ws["A6"].font = header_font
    ws["B6"] = str(data.get("time", "")) + " sec"

    ws["A7"] = "θ(t) Function"
    ws["A7"].font = header_font
    ws["B7"] = data.get("func", "")

    ws["A8"] = "Spring Constant"
    ws["A8"].font = header_font
    ws["B8"] = data.get("mode_name", "")

    # Bold left labels
    # for row in range(2, 9):
    #     ws[f"A{row}"].font = header_font

    for col_num in range(1, ws.max_column + 1):

        max_length = 0

        for row in range(2, ws.max_row + 1):  # Row 1 ignore

            value = ws.cell(row=row, column=col_num).value

            if value:
                max_length = max(max_length, len(str(value)))

        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    # =========================
    # TABLE HEADER
    # =========================

    ws["A10"] = "Gripper"
    ws["B10"] = "A"
    ws["C10"] = "K"
    ws["D10"] = "B"
    ws["E10"] = "F(t) (N)"

    for cell in ["A10", "B10", "C10", "D10", "E10"]:

        ws[cell].font = header_font
        ws[cell].alignment = Alignment(horizontal="center")

    # =========================
    # DATA ROWS
    # =========================

    rows = [
        ["Finger 1", data.get("a1"), data.get("k1"), data.get("b1"), data.get("f1")],
        ["Finger 2", data.get("a2"), data.get("k2"), data.get("b2"), data.get("f2")],
        ["Finger 3", data.get("a3"), data.get("k3"), data.get("b3"), data.get("f3")],
    ]

    # =========================
    # 4 Finger Gripper
    # =========================

    if "4" in str(data.get("gripper_name", "")):

        rows.append(
            ["Finger 4", data.get("a4"), data.get("k4"), data.get("b4"), data.get("f4")]
        )

    # =========================
    # 3 Finger + Thumb
    # =========================

    if "Thumb" in str(data.get("gripper_name", "")):

        rows.append(
            ["Thumb", data.get("ta"), data.get("tk"), data.get("tb"), data.get("ft")]
        )

    start_row = 9

    for row_data in rows:

        ws.append(row_data)

    # =========================
    # TOTAL
    # =========================

    ws["A15"] = "TOTAL"
    ws["E15"] = data.get("total")

    ws["A15"].font = Font(bold=True)
    ws["E15"].font = Font(bold=True)

    # =========================
    # COLUMN WIDTH
    # =========================
    # =========================
    # COLUMN WIDTH
    # =========================

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 12
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 12
    ws.column_dimensions["G"].width = 12
    ws.column_dimensions["H"].width = 12
    ws.column_dimensions["I"].width = 12
    ws.column_dimensions["J"].width = 12

    # for col_num, column_cells in enumerate(ws.columns, 1):

    #     length = max(len(str(cell.value or "")) for cell in column_cells)

    #     column_letter = get_column_letter(col_num)

    #     ws.column_dimensions[column_letter].width = length + 5

    # =========================================
    # DYNAMIC FILE NAME
    # =========================================

    current_time = datetime.now().strftime("%H-%M-%S")

    gripper_name = data.get("gripper_name", "").replace(" ", "").replace("+", "")

    shape_name = data.get("shape_name", "").replace(" ", "")

    material = data.get("material", "").replace(" ", "")

    file_name = (
        f"Input_"
        f"{gripper_name}_"
        f"{shape_name}_"
        f"{material}_"
        f"{current_time}.xlsx"
    )
    # =========================
    # SAVE FILE
    # =========================

    # file_name = "gripper_results.xlsx"
    file_path = os.path.join(UPLOAD_FOLDER, file_name)

    # wb.save(file_name)
    wb.save(file_path)

    # download_excel
    mStatus = data.get("mStatus")
    # print("mStatus::",mStatus)
    # return send_file(file_name, as_attachment=True, download_name=file_name)
    if mStatus == 1:
        # print("mStatus1::",mStatus)
        return send_file(file_path, as_attachment=True, download_name=file_name)
    else:
        # print("mStatus2::",mStatus)
        return jsonify(
            {
                "status": "success",
                "filename": file_name,
                "filepath": f"static/files/{file_name}",
                "message": f"{file_name} generated successfully",
            }
        )


@app.route("/download_excel", methods=["POST"])
def download_excel():

    data = request.json
    wb = Workbook()
    ws = wb.active

    if ws is None:
        raise Exception("Worksheet not created")

    ws.title = "Gripper Results"

    # =========================
    # HEADER DETAILS
    # =========================

    header_font = Font(bold=True)
    Grippertype = data.get("gripper_name", "")
    ws.merge_cells("A1:G1")
    ws["A1"] = "Schematic Representation of " f"Spring Structure Model ({Grippertype})"
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center")
    # ws.column_dimensions["A"].width = 15
    ws["A3"] = "Gripper type:"
    ws["A3"].font = header_font
    ws["B3"] = Grippertype

    ws["A4"] = "Object Shape"
    ws["A4"].font = header_font
    ws["B4"] = data.get("shape_name", "")

    ws["A5"] = "Material"
    ws["A5"].font = header_font
    ws["B5"] = data.get("material", "")

    ws["A6"] = "Time"
    ws["A6"].font = header_font
    ws["B6"] = str(data.get("time", "")) + " sec"

    ws["A7"] = "θ(t) Function"
    ws["A7"].font = header_font
    ws["B7"] = data.get("func", "")

    ws["A8"] = "Spring Constant"
    ws["A8"].font = header_font
    ws["B8"] = data.get("mode_name", "")

    ws["A10"] = "Volume (m³)"
    ws["A10"].font = header_font
    ws["B10"] = data.get("volume", "")

    ws["C10"] = "Mass (kg)"
    ws["C10"].font = header_font
    ws["D10"] = data.get("mass", "")

    # Bold left labels
    # for row in range(2, 9):
    #     ws[f"A{row}"].font = header_font

    for col_num in range(1, ws.max_column + 1):

        max_length = 0
        for row in range(2, ws.max_row + 1):  # Row 1 ignore
            value = ws.cell(row=row, column=col_num).value
            if value:
                max_length = max(max_length, len(str(value)))

        ws.column_dimensions[get_column_letter(col_num)].width = max_length + 2

    # =========================
    # TABLE HEADER
    # =========================
    start_header_row = 12
    # Main merged headers
    ws.merge_cells(f"A{start_header_row}:A{start_header_row+2}")
    ws.merge_cells(f"B{start_header_row}:C{start_header_row}")
    ws.merge_cells(f"D{start_header_row}:D{start_header_row}")
    ws.merge_cells(f"E{start_header_row}:F{start_header_row}")
    ws.merge_cells(f"G{start_header_row}:G{start_header_row+2}")

    # Row 10
    ws[f"A{start_header_row}"] = "Gripper"
    ws[f"B{start_header_row}"] = "|A|"
    ws[f"D{start_header_row}"] = "+K"
    ws[f"E{start_header_row}"] = "*|B|"
    ws[f"G{start_header_row}"] = "F(t) (N)"

    # Row 11
    ws[f"B{start_header_row+1}"] = "A1"
    ws[f"C{start_header_row+1}"] = "A2"

    ws[f"D{start_header_row+1}"] = "K"

    ws[f"E{start_header_row+1}"] = "B1"
    ws[f"F{start_header_row+1}"] = "B2"

    # Row 12
    ws[f"B{start_header_row+2}"] = "(mg)i"

    ws[f"C{start_header_row+2}"] = "[-ri*sin(θi)*(dθi/dt) + " "ri*cos(θi)*(d²θi/dt²)]"
    ws[f"D{start_header_row+2}"] = "(K1*K2)/(K1+K2)"

    ws[f"E{start_header_row+2}"] = "ri"
    ws[f"F{start_header_row+2}"] = "{θi - θi³/3! + θi⁵/5! + ...}"

    # =====================================================
    # STYLING
    # =====================================================

    header_font = Font(bold=True)
    thin = Side(style="thin")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    header_fill = PatternFill(
        start_color="D9D9D9", end_color="D9D9D9", fill_type="solid"
    )

    for row in ws.iter_rows(
        min_row=start_header_row, max_row=start_header_row + 2, min_col=1, max_col=7
    ):
        for cell in row:
            cell.font = header_font
            cell.alignment = Alignment(
                horizontal="center", vertical="center", wrap_text=True
            )
            cell.border = border
            cell.fill = header_fill

    # =====================================================
    # DATA ROWS
    # =====================================================

    data_row = start_header_row + 3
    rows = [
        [
            "Finger 1",
            data.get("a11"),
            data.get("a21"),
            data.get("k1"),
            data.get("b11"),
            data.get("b21"),
            data.get("f1"),
        ],
        [
            "Finger 2",
            data.get("a12"),
            data.get("a22"),
            data.get("k2"),
            data.get("b12"),
            data.get("b22"),
            data.get("f2"),
        ],
        [
            "Finger 3",
            data.get("a13"),
            data.get("a23"),
            data.get("k3"),
            data.get("b13"),
            data.get("b23"),
            data.get("f3"),
        ],
    ]
    # Finger 4
    if "4" in str(data.get("gripper_name", "")):
        rows.append(
            [
                "Finger 4",
                data.get("a14"),
                data.get("a24"),
                data.get("k4"),
                data.get("b14"),
                data.get("b24"),
                data.get("f4"),
            ]
        )

    # Thumb
    if "Thumb" in str(data.get("gripper_name", "")):
        rows.append(
            [
                "Thumb",
                data.get("a15"),
                data.get("ta"),
                data.get("tk"),
                data.get("b15"),
                data.get("tb"),
                data.get("ft"),
            ]
        )

    # Write rows
    for row_data in rows:
        ws.append(row_data)

    # =====================================================
    # TOTAL
    # =====================================================

    total_row = data_row + len(rows)
    ws[f"A{total_row}"] = "TOTAL"
    ws[f"G{total_row}"] = data.get("total", 0)

    ws[f"A{total_row}"].font = Font(bold=True)
    ws[f"G{total_row}"].font = Font(bold=True)

    # =====================================================
    # BORDERS FOR DATA
    # =====================================================

    for row in ws.iter_rows(min_row=data_row, max_row=total_row, min_col=1, max_col=7):
        for cell in row:
            cell.border = border
            cell.alignment = Alignment(horizontal="center", vertical="center")

    # =====================================================
    # COLUMN WIDTHS
    # =====================================================

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 12
    ws.column_dimensions["C"].width = 25
    ws.column_dimensions["D"].width = 18
    ws.column_dimensions["E"].width = 12
    ws.column_dimensions["F"].width = 25
    ws.column_dimensions["G"].width = 12

    # =========================================
    # DYNAMIC FILE NAME
    # =========================================

    current_time = datetime.now().strftime("%H-%M-%S")
    gripper_name = data.get("gripper_name", "").replace(" ", "").replace("+", "")
    shape_name = data.get("shape_name", "").replace(" ", "")
    material = data.get("material", "").replace(" ", "")

    file_name = (
        f"Input_"
        f"{gripper_name}_"
        f"{shape_name}_"
        f"{material}_"
        f"{current_time}.xlsx"
    )
    # =========================
    # SAVE FILE
    # =========================

    # file_name = "gripper_results.xlsx"
    file_path = os.path.join(UPLOAD_FOLDER, file_name)

    # wb.save(file_name)
    wb.save(file_path)

    # download_excel
    mStatus = data.get("mStatus")
    # print("mStatus::",mStatus)
    # return send_file(file_name, as_attachment=True, download_name=file_name)
    if mStatus == 1:
        # print("mStatus1::",mStatus)
        return send_file(file_path, as_attachment=True, download_name=file_name)
    else:
        # print("mStatus2::",mStatus)
        return jsonify(
            {
                "status": "success",
                "filename": file_name,
                "filepath": f"static/files/{file_name}",
                "message": f"{file_name} generated successfully",
            }
        )


@app.route("/download_results_pdf", methods=["POST"])
def download_results_pdf():

    data = request.json
    buffer = BytesIO()

    # doc = SimpleDocTemplate(buffer, pagesize=letter)
    doc = SimpleDocTemplate(buffer, pagesize=landscape(letter))
    styles = getSampleStyleSheet()
    elements = []

    # =====================================
    # TITLE
    # =====================================

    elements.append(
        Paragraph(
            "<b>Schematic Representation of Spring Structure Model for</b><br/>"
            f"{data.get('gripper_name')}",
            styles["Title"],
        )
    )

    elements.append(Spacer(1, 20))

    # =====================================
    # DETAILS
    # =====================================

    details = [
        f"<b>Gripper Type:</b> " f"{data.get('gripper_name')}",
        f"<b>Object Shape:</b> " f"{data.get('shape_name')}",
        f"<b>Material:</b> " f"{data.get('material')}",
        f"<b>Time:</b> " f"{data.get('time')} sec",
        f"<b>θ(t) Function:</b> " f"{data.get('func')}",
        f"<b>Spring Constant:</b> " f"{data.get('mode_name')}",
        f"<b>Volume (m³):</b> " f"{data.get('volume')}",
        f"<b>Mass (kg):</b> " f"{data.get('mass')}",
    ]

    for item in details:
        elements.append(Paragraph(item, styles["Normal"]))

    elements.append(Spacer(1, 20))

    # =====================================
    # TABLE
    # =====================================

    table_data = [
        ["", "|A|", "|A|", "+K", "*|B|", "*|B|", ""],
        ["Gripper", "A1", "A2", "K", "B1", "B2", "F(t) (N)"],
        [
            "",
            "(mg)i",
            "[-ri*sin(θi)*(dθi/dt)+ri*cos(θi)*(d²θi/dt²)]",
            "(K1*K2)/(K1+K2)",
            "ri",
            "{θi-θi³/3!+θi⁵/5!+...}",
            "",
        ],
    ]

    rows = [
        [
            "Finger 1",
            data.get("a11"),
            data.get("a21"),
            data.get("k1"),
            data.get("b11"),
            data.get("b21"),
            data.get("f1"),
        ],
        [
            "Finger 2",
            data.get("a12"),
            data.get("a22"),
            data.get("k2"),
            data.get("b12"),
            data.get("b22"),
            data.get("f2"),
        ],
        [
            "Finger 3",
            data.get("a13"),
            data.get("a23"),
            data.get("k3"),
            data.get("b13"),
            data.get("b23"),
            data.get("f3"),
        ],
    ]

    # Finger 4
    if "4" in str(data.get("gripper_name", "")):
        rows.append(
            [
                "Finger 4",
                data.get("a14"),
                data.get("a24"),
                data.get("k4"),
                data.get("b14"),
                data.get("b24"),
                data.get("f4"),
            ]
        )

    # Thumb
    if "Thumb" in str(data.get("gripper_name", "")):
        rows.append(
            [
                "Thumb",
                data.get("a15"),
                data.get("ta"),
                data.get("tk"),
                data.get("b15"),
                data.get("tb"),
                data.get("ft"),
            ]
        )

    table_data.extend(rows)

    table_data.append(
        [
            "TOTAL",
            "",
            "",
            "",
            "",
            "",
            data.get("total"),
        ]
    )

    table = Table(table_data, colWidths=[80, 60, 160, 80, 60, 100, 65])
    # table = Table(table_data, colWidths=[80, 70, 180, 90, 70, 180, 80])

    table.setStyle(
        TableStyle(
            [
                # Merge top headers
                ("SPAN", (1, 0), (2, 0)),  # |A|
                ("SPAN", (4, 0), (5, 0)),  # *|B|
                # Gripper span
                ("SPAN", (0, 1), (0, 2)),
                # F(t) span
                ("SPAN", (6, 1), (6, 2)),
                ("BACKGROUND", (0, 0), (-1, 2), colors.lightgrey),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                ("FONTNAME", (0, 0), (-1, 2), "Arial-Bold"),
                ("FONTNAME", (0, -1), (-1, -1), "Arial-Bold"),
                ("BACKGROUND", (0, -1), (-1, -1), colors.lightgrey),
                ("FONTSIZE", (0, 0), (-1, -1), 8),
            ]
        )
    )

    elements.append(table)

    # =====================================
    # BUILD PDF
    # =====================================

    doc.build(elements)

    buffer.seek(0)
    from datetime import datetime

    current_time = datetime.now().strftime("%H-%M-%S")
    gripper_name = data.get("gripper_name", "").replace(" ", "").replace("+", "")
    shape_name = data.get("shape_name", "").replace(" ", "")
    material = data.get("material", "").replace(" ", "")

    filename = (
        f"Input_"
        f"{gripper_name}_"
        f"{shape_name}_"
        f"{material}_"
        f"{current_time}.pdf"
    )

    return send_file(
        buffer, as_attachment=True, download_name=filename, mimetype="application/pdf"
    )


# CALL NEW API FROM graph.js
# CREATE COMMON FUNCTION
def perform_calculation(data, t):

    shape = int(data.get("shape", 0))
    event = data["event"]

    length = float(data.get("length", 0)) * 0.001
    breadth = float(data.get("breadth", 0)) * 0.001
    width = float(data.get("width", 0)) * 0.001

    radius = float(data.get("radius", 0)) * 0.001

    Rmajor = float(data.get("Rmajor", 0)) * 0.001
    Rminor = float(data.get("Rminor", 0)) * 0.001

    material = data["material"]

    func = data["func"]

    gripper = int(data["gripper"])
    # gripper = int(data["gripper"])
    # print("Gripper type:", gripper)

    # =====================================
    # VOLUME
    # =====================================

    if shape == 1:

        if event == "length":
            L, B, H = length, breadth, width

        elif event == "breadth":
            L, B, H = breadth, length, width

        else:
            L, B, H = 0.1, 0.04, 0.02

        volume = L * B * H

    elif shape == 2:

        volume = (4 / 3) * math.pi * (radius**3)

    elif shape == 3:

        if event == "major":
            a, b, c = Rmajor, Rminor, Rminor

        elif event == "minor":
            a, b, c = Rminor, Rminor, Rmajor

        else:
            a, b, c = 0.05, 0.03, 0.03

        volume = (4 / 3) * math.pi * a * b * c

    else:
        volume = 0

    density = materials.get(material, 0)

    M = density * volume

    forces = []

    thumb = 0

    finger_count = 4 if gripper == 1 else 3

    mode = data["mode"]

    # =====================================
    # MODE 1
    # =====================================

    if mode == "1":

        k = float(data.get("k_common", 0))

        kf = (k * k) / (k + k)

        for i in range(finger_count):

            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, kf, func, t)

            F = abs(A) + K * abs(B)

            forces.append(round(F, 4))

        if gripper == 2:

            ktt = 1 / ((1 / k) + (1 / k) + (1 / k))

            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, ktt, func, t)

            thumb = round(abs(A) + K * abs(B), 4)

    # =====================================
    # MODE 2
    # =====================================

    elif mode == "2":

        kf = float(data.get("k_finger", 0))

        kf = (kf * kf) / (kf + kf)

        for i in range(finger_count):

            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, kf, func, t)

            F = abs(A) + K * abs(B)

            forces.append(round(F, 4))

        if gripper == 2:

            kt = data.get("thumb", [])

            if len(kt) >= 3:

                ktt = 1 / ((1 / kt[0]) + (1 / kt[1]) + (1 / kt[2]))

                A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, ktt, func, t)

                thumb = round(abs(A) + K * abs(B), 4)

    # =====================================
    # MODE 3
    # =====================================

    elif mode == "3":

        for i in range(finger_count):

            k1 = data["fingers"][i]["k1"]

            k2 = data["fingers"][i]["k2"]

            kf = (k1 * k2) / (k1 + k2)

            A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, kf, func, t)

            F = abs(A) + K * abs(B)

            forces.append(round(F, 4))

        if gripper == 2:

            kt = data.get("thumb", [])

            if len(kt) >= 3:

                ktt = 1 / ((1 / kt[0]) + (1 / kt[1]) + (1 / kt[2]))

                A1, A2, B1, B2, A, B, K = cal_force_eq14A1B1(M, ktt, func, t)

                thumb = round(abs(A) + K * abs(B), 4)

    total = round(sum(forces) + thumb, 4)
    # print(f"t={t} sec, Forces: {forces}, Thumb: {thumb}, Total: {total}")

    return total


@app.route("/calculate_graph", methods=["POST"])
def calculate_graph():
    start_time = time.perf_counter()
    data = request.json
    time_data = []
    force_data = []

    # print("Data received:", data)
    # Get All previous entries using get_saved_input_graph_all function and according to recieve how many time_value we have to calculate, we will filter the data and send to frontend for graph plotting
    previous_entries = get_saved_input_graph_all(
        gripper=data.get("gripper"),
        shape=data.get("shape"),
        material=data.get("material"),
        theta_function=data.get("func"),
        spring_mode=data.get("mode"),
    )
    # print("Data received previous_entries:", previous_entries)
    # for t in range(1, 6):
    #     total_force = perform_calculation(data, t)
    #     time_data.append(f"{t} sec")
    #     force_data.append(total_force)

    for entry in previous_entries:

        t = float(entry["time_value"])
        total_force = perform_calculation(data, t)
        time_data.append(f"{t:g} sec")
        force_data.append(total_force)

    end_time = time.perf_counter()
    execution_time = (end_time - start_time) * 1_000_000

    return jsonify(
        {
            "time": time_data,
            "force": force_data,
            "execution_time": round(execution_time, 2),
        }
    )


@app.route("/download_graph_excel", methods=["POST"])
def download_graph_excel():

    data = request.json

    wb = Workbook()

    ws = wb.active

    if ws is None:
        raise Exception("Worksheet not created")

    ws.title = "Graph Report"

    # =========================================
    # MAIN HEADING
    # =========================================

    ws.merge_cells("A1:J1")

    # It is gripper name:
    shape_name = data.get("shape_name", "")
    ws["A1"] = "Bar Chart Representation of " f"Spring Structure Model ({shape_name})"
    ws["A1"].font = Font(bold=True, size=18)

    ws["A1"].alignment = Alignment(horizontal="center")

    # =========================================
    # OBJECT DETAILS
    # =========================================

    object_shape = data.get("object_shape", "-")

    material = data.get("material", "-")

    ws["A4"] = f"Object Shape : {object_shape}"

    ws["A5"] = f"Material : {material}"

    ws["A3"].font = Font(bold=True, size=12)

    ws["A4"].font = Font(bold=True, size=12)
    ws["A3"] = f"Gripper Type : " f"{data.get('gripper_name', '-')}"

    ws["A6"] = f"θ(t) Function : " f"{data.get('func', '-')}"

    ws["A7"] = f"Spring Constant : " f"{data.get('mode_name', '-')}"

    ws["A5"].font = Font(bold=True, size=12)

    ws["A6"].font = Font(bold=True, size=12)

    ws["A7"].font = Font(bold=True, size=12)
    # =========================================
    # TABLE HEADER
    # =========================================

    ws["A9"] = "Time t (sec)"
    ws["B9"] = "Force (N)"

    # from openpyxl.styles import Alignment
    ws["B9"].alignment = Alignment(horizontal="right")

    header_font = Font(bold=True)

    ws["A9"].font = header_font
    ws["B9"].font = header_font

    # =========================================
    # TABLE DATA
    # =========================================

    row = 10
    for item in data["tableData"]:

        ws[f"A{row}"] = item["time"]

        # ws[f"B{row}"] = float(item["force"])
        force_value = str(item["force"]).replace("(N)", "").strip()
        ws[f"B{row}"] = float(force_value)

        # print("Time:", item["time"], "Force:", item["force"])

        row += 1

    # =========================================
    # COLUMN WIDTH
    # =========================================

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20

    # =========================================
    # GRAPH IMAGE
    # =========================================

    # graph_data = data["graphImage"]
    graph_data = data.get("graphImage")
    if not graph_data:
        return jsonify({"error": "Graph image missing"}), 400

    if not graph_data:
        return jsonify({"error": "Graph image not received"}), 400

    graph_data = graph_data.split(",")[1]

    image_data = base64.b64decode(graph_data)

    image_stream = BytesIO(image_data)

    img = Image(image_stream)

    img.width = 900
    img.height = 450

    ws.add_image(img, "A16")

    # =========================================
    # SAVE FILE
    # =========================================
    shape_name = shape_name.replace(" ", "").replace("+", "")
    # filename = "Graph_Report_AAAA.xlsx"
    current_time = datetime.now().strftime("%H-%M-%S")

    filename = (
        f"Graph_" f"{shape_name}_" f"{object_shape}_" f"{material}_{current_time}.xlsx"
    )

    # print("Generating file:", filename)

    # excel_file = BytesIO()

    # wb.save(excel_file)

    # excel_file.seek(0)

    # return send_file(
    #     excel_file,
    #     as_attachment=True,
    #     download_name=filename,
    #     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    # )
    file_path = os.path.join(UPLOAD_FOLDER, filename)

    wb.save(file_path)

    mStatus = data.get("mStatus")
    if mStatus == 1:
        # print("mStatus1::",mStatus)
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        # print("mStatus2::",mStatus)
        return jsonify(
            {
                "status": "success",
                "filename": filename,
                "filepath": f"static/files/{filename}",
                "message": f"{filename} generated successfully",
            }
        )


@app.route("/download_compare_excel", methods=["POST"])
def download_compare_excel():

    data = request.json
    wb = Workbook()
    ws = wb.active

    if ws is None:
        raise Exception("Worksheet not created")

    ws.title = "Comparison Chart"

    # =========================================
    # MAIN HEADING
    # =========================================

    ws.merge_cells("A1:K1")
    shape_name = data.get("shape_name", "")
    ws["A1"] = (
        "Comparison Chart Representation of " f"Spring Structure Model ({shape_name})"
    )
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center")

    # =========================================
    # OBJECT DETAILS
    # =========================================

    object_shape = data.get("object_shape", "-")
    time_val = data.get("time", "-")
    ws["A4"] = f"Object Shape : {object_shape}"
    ws["A5"] = f"Time : {time_val}"
    ws["A3"].font = Font(bold=True, size=12)
    ws["A4"].font = Font(bold=True, size=12)
    ws["A3"] = f"Gripper Type : " f"{data.get('gripper_name', '-')}"
    ws["A6"] = f"θ(t) Function : " f"{data.get('func', '-')}"
    ws["A7"] = f"Spring Constant : " f"{data.get('mode_name', '-')}"
    ws["A5"].font = Font(bold=True, size=12)
    ws["A6"].font = Font(bold=True, size=12)
    ws["A7"].font = Font(bold=True, size=12)
    # =========================================
    # TABLE HEADER
    # =========================================
    tab = data.get("tab", "0")
    if tab == 1:
        ws["A9"] = "Materials"
    else:
        ws["A9"] = "All equal"
        # ws["A9"].alignment = Alignment(horizontal="right")

    ws["B9"] = "Force (N)"
    # from openpyxl.styles import Alignment
    ws["B9"].alignment = Alignment(horizontal="right")
    header_font = Font(bold=True)
    ws["A9"].font = header_font
    ws["B9"].font = header_font

    # =========================================
    # TABLE DATA
    # =========================================

    row = 10
    for item in data["tableData"]:

        ws[f"A{row}"] = item["time"]
        # ws[f"B{row}"] = float(item["force"])
        force_value = str(item["force"]).replace("(N)", "").strip()
        ws[f"B{row}"] = float(force_value)
        # print("Time:", item["time"], "Force:", item["force"])
        row += 1

    # =========================================
    # COLUMN WIDTH
    # =========================================

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20

    # =========================================
    # GRAPH IMAGE
    # =========================================
    # graph_data = data["graphImage"]
    graph_data = data.get("graphImage")

    if not graph_data:
        return jsonify({"error": "Graph image not received"}), 400

    graph_data = graph_data.split(",")[1]
    image_data = base64.b64decode(graph_data)
    image_stream = BytesIO(image_data)
    img = Image(image_stream)
    img.width = 900
    img.height = 450

    ws.add_image(img, "A16")

    # =========================================
    # SAVE FILE
    # =========================================
    shape_name = shape_name.replace(" ", "").replace("+", "")
    # filename = "Graph_Report_AAAA.xlsx"
    current_time = datetime.now().strftime("%H-%M-%S")

    filename = (
        f"Graph_Comparison_"
        f"{shape_name}_"
        f"{object_shape}_"
        f"{time_val}_{current_time}.xlsx"
    )

    file_path = os.path.join(UPLOAD_FOLDER, filename)
    wb.save(file_path)
    mStatus = data.get("mStatus")
    if mStatus == 1:
        # print("mStatus1::",mStatus)
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        # print("mStatus2::",mStatus)
        return jsonify(
            {
                "status": "success",
                "filename": filename,
                "filepath": f"static/files/{filename}",
                "message": f"{filename} generated successfully",
            }
        )


# =================== tab 3 =====================
@app.route("/download_compare_excel2", methods=["POST"])
def download_compare_excel2():

    data = request.json
    wb = Workbook()
    ws = wb.active

    if ws is None:
        raise Exception("Worksheet not created")

    ws.title = "Comparison Chart"

    # =========================================
    # MAIN HEADING
    # =========================================

    ws.merge_cells("A1:K1")
    shape_name = data.get("shape_name", "")
    ws["A1"] = (
        "Comparison Chart Representation of " f"Spring Structure Model ({shape_name})"
    )
    ws["A1"].font = Font(bold=True, size=18)
    ws["A1"].alignment = Alignment(horizontal="center")

    # =========================================
    # OBJECT DETAILS
    # =========================================

    material = data.get("material", "-")
    time_val = data.get("time", "-")
    ws["A4"] = f"Material : {material}"
    ws["A5"] = f"Time : {time_val}"
    ws["A3"].font = Font(bold=True, size=12)
    ws["A4"].font = Font(bold=True, size=12)
    ws["A3"] = f"Gripper Type : " f"{data.get('gripper_name', '-')}"
    ws["A6"] = f"θ(t) Function : " f"{data.get('func', '-')}"
    ws["A7"] = f"Spring Constant : " f"{data.get('mode_name', '-')}"
    ws["A5"].font = Font(bold=True, size=12)
    ws["A6"].font = Font(bold=True, size=12)
    ws["A7"].font = Font(bold=True, size=12)
    # =========================================
    # TABLE HEADER
    # =========================================
    tab = data.get("tab", "0")
    if tab == 1:
        ws["A9"] = "Object Shape"
    else:
        ws["A9"] = "-"
        # ws["A9"].alignment = Alignment(horizontal="right")

    ws["B9"] = "Force (N)"
    # from openpyxl.styles import Alignment
    ws["B9"].alignment = Alignment(horizontal="right")
    header_font = Font(bold=True)
    ws["A9"].font = header_font
    ws["B9"].font = header_font

    # =========================================
    # TABLE DATA
    # =========================================

    row = 10
    for item in data["tableData"]:

        ws[f"A{row}"] = item["time"]
        # ws[f"B{row}"] = float(item["force"])
        force_value = str(item["force"]).replace("(N)", "").strip()
        ws[f"B{row}"] = float(force_value)
        # print("Time:", item["time"], "Force:", item["force"])
        row += 1

    # =========================================
    # COLUMN WIDTH
    # =========================================

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 20

    # =========================================
    # GRAPH IMAGE
    # =========================================
    # graph_data = data["graphImage"]
    graph_data = data.get("graphImage")

    if not graph_data:
        return jsonify({"error": "Graph image not received"}), 400

    graph_data = graph_data.split(",")[1]
    image_data = base64.b64decode(graph_data)
    image_stream = BytesIO(image_data)
    img = Image(image_stream)
    img.width = 900
    img.height = 450

    ws.add_image(img, "A16")

    # =========================================
    # SAVE FILE
    # =========================================
    shape_name = shape_name.replace(" ", "").replace("+", "")
    # filename = "Graph_Report_AAAA.xlsx"
    current_time = datetime.now().strftime("%H-%M-%S")

    filename = (
        f"Graph_Comparison_"
        f"{shape_name}_"
        f"{material}_"
        f"{time_val}_{current_time}.xlsx"
    )

    file_path = os.path.join(UPLOAD_FOLDER, filename)
    wb.save(file_path)
    mStatus = data.get("mStatus")
    if mStatus == 1:
        # print("mStatus1::",mStatus)
        return send_file(
            file_path,
            as_attachment=True,
            download_name=filename,
            mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )
    else:
        # print("mStatus2::",mStatus)
        return jsonify(
            {
                "status": "success",
                "filename": filename,
                "filepath": f"static/files/{filename}",
                "message": f"{filename} generated successfully",
            }
        )


# ==================== End tab 3 ======================


@app.route("/download_graph_pdf", methods=["POST"])
def download_graph_pdf():

    data = request.json

    buffer = BytesIO()

    doc = SimpleDocTemplate(buffer, pagesize=letter)

    elements = []

    styles = getSampleStyleSheet()

    # =========================================
    # TITLE
    # =========================================

    shape_name = data.get("shape_name", "")

    title = "Bar Chart Representation of " f"Spring Structure Model " f"({shape_name})"

    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))

    elements.append(Spacer(1, 20))

    # =========================================
    # DETAILS
    # =========================================

    object_shape = data.get("object_shape", "-")

    material = data.get("material", "-")

    elements.append(
        Paragraph(
            f"<b>Gripper Type:</b> " f"{data.get('gripper_name', '-')}",
            styles["Normal"],
        )
    )

    elements.append(Paragraph(f"<b>Object Shape:</b> {object_shape}", styles["Normal"]))

    elements.append(Paragraph(f"<b>Material:</b> {material}", styles["Normal"]))

    elements.append(
        Paragraph(
            f"<b>θ(t) Function:</b> " f"{data.get('func', '-')}", styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Spring Constant:</b> " f"{data.get('mode_name', '-')}",
            styles["Normal"],
        )
    )

    elements.append(Spacer(1, 20))

    # =========================================
    # TABLE
    # =========================================

    table_data = [["Time t (sec)", "Force (N)"]]

    for item in data["tableData"]:

        table_data.append([item["time"], item["force"]])

    table = Table(table_data, colWidths=[200, 200])

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(table)

    elements.append(Spacer(1, 30))

    # =========================================
    # GRAPH IMAGE
    # =========================================

    graph_data = data["graphImage"]

    graph_data = graph_data.split(",")[1]

    image_data = base64.b64decode(graph_data)

    temp_image = "temp_chart_pdf.png"

    with open(temp_image, "wb") as f:

        f.write(image_data)

    chart = RLImage(temp_image, width=500, height=280)

    elements.append(chart)

    # =========================================
    # BUILD PDF
    # =========================================

    doc.build(elements)

    buffer.seek(0)

    # delete temp image
    if os.path.exists(temp_image):

        os.remove(temp_image)

    # =========================================
    # DOWNLOAD
    # =========================================
    current_time = datetime.now().strftime("%H-%M-%S")

    filename = (
        f"Graph_"
        f"{shape_name}_"
        f"{object_shape}_"
        f"{material}_"
        f"{current_time}.pdf"
    )

    return send_file(
        buffer, as_attachment=True, download_name=filename, mimetype="application/pdf"
    )


@app.route("/download_compare_pdf", methods=["POST"])
def download_compare_pdf():

    data = request.json
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # =========================================
    # TITLE
    # =========================================

    shape_name = data.get("shape_name", "")
    title = (
        "Comparison Chart Representation of "
        f"Spring Structure Model "
        f"({shape_name})"
    )

    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Spacer(1, 20))

    # =========================================
    # DETAILS
    # =========================================

    object_shape = data.get("object_shape", "-")
    time_val = data.get("time", "-")
    elements.append(
        Paragraph(
            f"<b>Gripper Type:</b> " f"{data.get('gripper_name', '-')}",
            styles["Normal"],
        )
    )

    elements.append(Paragraph(f"<b>Object Shape:</b> {object_shape}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Time:</b> {time_val}", styles["Normal"]))
    elements.append(
        Paragraph(
            f"<b>θ(t) Function:</b> " f"{data.get('func', '-')}", styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Spring Constant:</b> " f"{data.get('mode_name', '-')}",
            styles["Normal"],
        )
    )

    elements.append(Spacer(1, 20))

    # =========================================
    # TABLE
    # =========================================
    tab = data.get("tab", "0")
    if tab == 1:
        table_data = [["Materials", "Force (N)"]]
    else:
        table_data = [["All equal", "Force (N)"]]

    for item in data["tableData"]:

        table_data.append([item["time"], item["force"]])

    table = Table(table_data, colWidths=[200, 200])

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(table)
    elements.append(Spacer(1, 30))

    # =========================================
    # GRAPH IMAGE
    # =========================================

    graph_data = data["graphImage"]
    graph_data = graph_data.split(",")[1]
    image_data = base64.b64decode(graph_data)
    temp_image = "temp_chart_pdf.png"

    with open(temp_image, "wb") as f:
        f.write(image_data)

    chart = RLImage(temp_image, width=500, height=280)
    elements.append(chart)

    # =========================================
    # BUILD PDF
    # =========================================

    doc.build(elements)
    buffer.seek(0)

    # delete temp image
    if os.path.exists(temp_image):
        os.remove(temp_image)

    # =========================================
    # DOWNLOAD
    # =========================================
    current_time = datetime.now().strftime("%H-%M-%S")

    filename = (
        f"Graph_Comparison_"
        f"{shape_name}_"
        f"{object_shape}_"
        f"{time_val}_"
        f"{current_time}.pdf"
    )

    return send_file(
        buffer, as_attachment=True, download_name=filename, mimetype="application/pdf"
    )


@app.route("/download_compare_pdf2", methods=["POST"])
def download_compare_pdf2():

    data = request.json
    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=letter)
    elements = []
    styles = getSampleStyleSheet()

    # =========================================
    # TITLE
    # =========================================

    shape_name = data.get("shape_name", "")
    title = (
        "Comparison Chart Representation of "
        f"Spring Structure Model "
        f"({shape_name})"
    )

    elements.append(Paragraph(f"<b>{title}</b>", styles["Title"]))
    elements.append(Spacer(1, 20))

    # =========================================
    # DETAILS
    # =========================================

    material = data.get("material", "-")
    time_val = data.get("time", "-")
    elements.append(
        Paragraph(
            f"<b>Gripper Type:</b> " f"{data.get('gripper_name', '-')}",
            styles["Normal"],
        )
    )

    elements.append(Paragraph(f"<b>Material:</b> {material}", styles["Normal"]))
    elements.append(Paragraph(f"<b>Time:</b> {time_val}", styles["Normal"]))
    elements.append(
        Paragraph(
            f"<b>θ(t) Function:</b> " f"{data.get('func', '-')}", styles["Normal"]
        )
    )

    elements.append(
        Paragraph(
            f"<b>Spring Constant:</b> " f"{data.get('mode_name', '-')}",
            styles["Normal"],
        )
    )

    elements.append(Spacer(1, 20))

    # =========================================
    # TABLE
    # =========================================
    tab = data.get("tab", "0")
    if tab == 1:
        table_data = [["Object Shape", "Force (N)"]]
    else:
        table_data = [["All equal", "Force (N)"]]

    for item in data["tableData"]:

        table_data.append([item["time"], item["force"]])

    table = Table(table_data, colWidths=[200, 200])

    table.setStyle(
        TableStyle(
            [
                ("BACKGROUND", (0, 0), (-1, 0), colors.grey),
                ("TEXTCOLOR", (0, 0), (-1, 0), colors.whitesmoke),
                ("ALIGN", (0, 0), (-1, -1), "CENTER"),
                ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                ("BOTTOMPADDING", (0, 0), (-1, 0), 12),
                ("GRID", (0, 0), (-1, -1), 1, colors.black),
            ]
        )
    )

    elements.append(table)
    elements.append(Spacer(1, 30))

    # =========================================
    # GRAPH IMAGE
    # =========================================

    graph_data = data["graphImage"]
    graph_data = graph_data.split(",")[1]
    image_data = base64.b64decode(graph_data)
    temp_image = "temp_chart_pdf.png"

    with open(temp_image, "wb") as f:
        f.write(image_data)

    chart = RLImage(temp_image, width=500, height=280)
    elements.append(chart)

    # =========================================
    # BUILD PDF
    # =========================================

    doc.build(elements)
    buffer.seek(0)

    # delete temp image
    if os.path.exists(temp_image):
        os.remove(temp_image)

    # =========================================
    # DOWNLOAD
    # =========================================
    current_time = datetime.now().strftime("%H-%M-%S")

    filename = (
        f"Graph_Comparison_"
        f"{shape_name}_"
        f"{material}_"
        f"{time_val}_"
        f"{current_time}.pdf"
    )

    return send_file(
        buffer, as_attachment=True, download_name=filename, mimetype="application/pdf"
    )


@app.route("/get_spring_constants", methods=["POST"])
def get_spring_constants_api():

    # print(">>>>>> get_spring_constants API called...........")
    data = request.json
    # print("Data received for spring constants:", data)

    gripper = int(data.get("gripper", 0))
    shape = int(data.get("shape", 0))
    material = data.get("material", "")
    # time_value = float(data.get("time", 0))
    theta_function = data.get("func", "")
    spring_key = data.get("spring_key", "")

    values = get_spring_constants(
        gripper,
        shape,
        material,
        # time_value,
        theta_function,
        spring_key,
    )

    # print("@@@@@ spring_values:", values)
    # print("Spring Key:", spring_key)

    # result = [
    #     row["spring_value"]
    #     for row in values
    # ]
    # # print("Result:", result)
    # return jsonify({
    #     "spring_values": result
    # })
    return jsonify({"spring_values": values})


@app.route("/comparison_data", methods=["POST"])
def comparison_data():
    # print("comparison_data API called")
    start = time.perf_counter()
    # existing code
    data = request.json
    # print("data>>", data)
    # materials = ["Rubber", "ABS", "Teflon"]
    materials = ["rubber", "abs", "teflon"]
    result = []
    for material in materials:
        temp = dict(data)
        temp["material"] = material
        total_force = perform_calculation(temp, float(data["time"]))

        if material == "rubber":
            material = "Rubber"
        elif material == "abs":
            material = "ABS"
        elif material == "teflon":
            material = "Teflon"
        # print("Material>>>", temp["material"])
        result.append({"material": material, "force": round(total_force, 4)})

    # print("Comparison Result:", result)
    elapsed_us = (time.perf_counter() - start) * 1_000_000  # Convert to microseconds
    # return jsonify(result)
    return jsonify({"result": result, "execution_time_us": round(elapsed_us, 2)})


# for tab 3
@app.route("/comparison_data2", methods=["POST"])
def comparison_data2():
    # print("comparison_data API called")
    start = time.perf_counter()
    # existing code
    data = request.json
    # materials = ["Rubber", "ABS", "Teflon"]
    shapes = [1, 2, 3]
    result = []
    for shape in shapes:
        temp = dict(data)
        temp["shape"] = shape
        total_force = perform_calculation(temp, float(data["time"]))
        if shape == 1:
            shape = "Rectangular Paralleopiped"
        elif shape == 2:
            shape = "Spherical"
        elif shape == 3:
            shape = "Ellipsoidal"

        # print("Material>>>", temp["material"])
        result.append({"shape": shape, "force": round(total_force, 4)})

    # print("Comparison Result3:", result)
    elapsed_us = (time.perf_counter() - start) * 1_000_000  # Convert to microseconds
    # return jsonify(result)
    return jsonify({"result": result, "execution_time_us": round(elapsed_us, 2)})


# for tab 2
@app.route("/comparison_data1", methods=["POST"])
def comparison_data1():
    # print("comparison_data1 API called")
    start = time.perf_counter()
    # existing code
    data = request.json
    gripper = int(data.get("gripper", 0))
    shape = int(data.get("shape", 0))
    material = data.get("material", "")
    theta_function = data.get("func", "")
    time_value = float(data.get("time", 0))

    # Fetch all Spring Constant (All equal) from spring_constants where spring_key = 'k_common'
    # print("gripper:", gripper)
    # print("shape:", shape)
    # print("material:", material)
    # print("theta_function:", theta_function)
    # print("time_value:", time_value)
    values = get_comparison_all_equal(
        gripper, shape, material, theta_function, time_value
    )
    # print("values:", values)
    k_common_all = [row["spring_value"] for row in values]
    # print("k_common_all>>>>" , k_common_all)

    result = []

    for k_common in k_common_all:
        temp = dict(data)
        temp["k_common"] = k_common
        # print("temp =", temp)
        total_force = perform_calculation(temp, float(data["time"]))
        result.append({"k_common": k_common, "force": round(total_force, 4)})

    # print("Comparison Result:", result)
    elapsed_us = (time.perf_counter() - start) * 1_000_000  # Convert to microseconds
    # return jsonify(result)
    return jsonify({"result": result, "execution_time_us": round(elapsed_us, 2)})


@app.route("/get_comparison_time", methods=["POST"])
def get_comparison_time_api():

    data = request.json

    gripper = int(data.get("gripper", 0))

    shape = int(data.get("shape", 0))

    theta_function = data.get("func", "")

    values = get_comparison_time(gripper, shape, theta_function)

    result = [row["time_value"] for row in values]

    return jsonify({"times": result})


# for tab 2
@app.route("/get_comparison_time1", methods=["POST"])
def get_comparison_time1_api():

    data = request.json
    gripper = int(data.get("gripper", 0))
    shape = int(data.get("shape", 0))
    material = data.get("material", 0)
    theta_function = data.get("func", "")

    values = get_comparison_time1(gripper, shape, material, theta_function)
    result = [row["time_value"] for row in values]

    return jsonify({"times": result})


# for tab 3
@app.route("/get_comparison_time2", methods=["POST"])
def get_comparison_time2_api():
    data = request.json
    gripper = int(data.get("gripper", 0))
    material = data.get("material", "")
    theta_function = data.get("func", "")
    values = get_comparison_time2(gripper, material, theta_function)
    result = [row["time_value"] for row in values]
    return jsonify({"times": result})


# for tab 1
@app.route("/get_spring_constants_comparison", methods=["POST"])
def get_spring_constants_comparison_api():

    data = request.json
    gripper = int(data.get("gripper", 0))
    shape = int(data.get("shape", 0))
    theta_function = data.get("func", "")
    time_value = float(data.get("time", 0))
    spring_key = data.get("spring_key", "")

    values = get_spring_constants_comparison(
        gripper, shape, theta_function, time_value, spring_key
    )
    # print("values>>", values)
    # result = [row["spring_value"] for row in values]
    result = [
        {"spring_value": row["spring_value"], "time_value": row["time_value"]}
        for row in values
    ]
    # print("Comparison Spring Constants:", result)
    return jsonify({"spring_values": result})


# for tab 3
@app.route("/get_spring_constants_comparison2", methods=["POST"])
def get_spring_constants_comparison2_api():

    data = request.json
    gripper = int(data.get("gripper", 0))
    material = data.get("material", "")
    theta_function = data.get("func", "")
    time_value = float(data.get("time", 0))
    spring_key = data.get("spring_key", "")

    # print(data)

    values = get_spring_constants_comparison2(
        gripper, material, theta_function, time_value, spring_key
    )
    # print("values>>", values)
    result = [row["spring_value"] for row in values]
    result = [
        {"spring_value": row["spring_value"], "time_value": row["time_value"]}
        for row in values
    ]
    # print("Comparison Spring Constants:", result)
    return jsonify({"spring_values": result})


@app.route("/merge_excels", methods=["POST"])
def merge_excels():

    data = request.json

    files = data.get("files", [])

    merge_list = []

    for item in files:

        filename = item["filename"]
        narration = item["narration"]

        file_path = os.path.join(UPLOAD_FOLDER, filename)

        merge_list.append((file_path, narration))

    # print(merge_list)
    # print("\nFiles to Merge:")

    # for filename, narration in merge_list:
    #     print(f"File: {filename}")
    #     print(f"Narration: {narration}")
    #     print("-" * 80)

    file_path = fileMerged(merge_list, UPLOAD_FOLDER)
    filename = "Final_Report_2026.xlsx"
    # return jsonify({
    #     "status": "success"
    # })

    return send_file(
        file_path,
        as_attachment=True,
        download_name=filename,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


# import webbrowser
# from threading import Timer

if __name__ == "__main__":
    # Timer(2, lambda: webbrowser.open("http://localhost:8000")).start()
    # app.run()
    # print("===>> Start App on", "http://127.0.0.1:8000")
    # waitress_serve(app, host="0.0.0.0", port=8000, threads=8)
    port = int(os.environ.get("PORT", 10000))
    print("===>> Start App on", f"http://127.0.0.1:{port}")
    # app.run(host="0.0.0.0", port=port, debug=True)
    # print(f"===>> Start App on port {port}")
    waitress_serve(app, host="0.0.0.0", port=port, threads=8)

# Version Control Commands (Git)
# git status
# git add .
# git commit -m "describe your changes"
# git commit -m "Updated index and app"
# # Pull latest (SAFE PRACTICE)
# git pull origin main --rebase
# git push
# git push -f origin main

# Faster Version (Important) Start Command::
# gunicorn app:app --bind 0.0.0.0:$PORT --workers 1 --threads 8

# Add the command in Render Dashboard → Service → Settings → Build Command:

# pip install --prefer-binary -r requirements.txt && python -c "import matplotlib.pyplot as plt"

# And Start Command:

# gunicorn --bind 0.0.0.0:$PORT app:app --workers 1 --threads 8

# This is the correct Render setup.
