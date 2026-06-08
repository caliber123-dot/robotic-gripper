
# ======================================
from copy import copy
from openpyxl import load_workbook, Workbook
from openpyxl.drawing.image import Image
import tempfile
import os
def copy_sheet(src_ws, dst_ws):

    # Copy cells + styles
    for row in src_ws.iter_rows():
        for cell in row:

            new_cell = dst_ws[cell.coordinate]
            new_cell.value = cell.value

            if cell.has_style:
                new_cell.font = copy(cell.font)
                new_cell.fill = copy(cell.fill)
                new_cell.border = copy(cell.border)
                new_cell.alignment = copy(cell.alignment)
                new_cell.number_format = copy(cell.number_format)
                new_cell.protection = copy(cell.protection)

    # Copy column widths
    for col_letter, col_dim in src_ws.column_dimensions.items():
        dst_ws.column_dimensions[col_letter].width = col_dim.width

    # Copy row heights
    for row_num, row_dim in src_ws.row_dimensions.items():
        dst_ws.row_dimensions[row_num].height = row_dim.height

    # Copy merged cells
    for merged_range in src_ws.merged_cells.ranges:
        dst_ws.merge_cells(str(merged_range))

    # Copy images
    for img in getattr(src_ws, "_images", []):
        try:
            image_bytes = img._data()

            tmp = tempfile.NamedTemporaryFile(
                delete=False,
                suffix=".png"
            )

            tmp.write(image_bytes)
            tmp.close()

            new_img = Image(tmp.name)

            try:
                anchor = img.anchor._from
                cell_ref = f"{chr(anchor.col + 65)}{anchor.row + 1}"
            except:
                cell_ref = "A1"

            dst_ws.add_image(new_img, cell_ref)

        except Exception as e:
            print("Image copy failed:", e)

# ----------------------------------
# Load source workbooks
# ----------------------------------

# wb1 = load_workbook("Input.xlsx")
# wb2 = load_workbook("Graph.xlsx")

# src1 = wb1.active
# src2 = wb2.active

# if src1 is None:
#     raise ValueError("No active sheet in Input.xlsx")

# if src2 is None:
#     raise ValueError("No active sheet in Graph.xlsx")

# # ----------------------------------
# # Create merged workbook
# # ----------------------------------

# new_wb = Workbook()

# ws1 = new_wb.active

# if ws1 is None:
#     raise ValueError("Failed to create worksheet")

# ws1.title = src1.title
# ws1.title = "Gripper Results (1)"

# # Copy first sheet
# copy_sheet(src1, ws1)

# # Create and copy second sheet
# # ws2 = new_wb.create_sheet(title=src2.title)
# ws2 = new_wb.create_sheet("Graph Comparison (1)")

# copy_sheet(src2, ws2)

# # ----------------------------------
# # Save
# # ----------------------------------

# new_wb.save("merged.xlsx")

# print("Merged successfully!")
# ----------------------------------
# Files to merge
# ----------------------------------

def fileMerged(files, UPLOAD_FOLDER):

    # files = [
    #     ("Input1.xlsx", "Input 1"),
    #     # ("Input2.xlsx", "Input 2"),
    #     ("Graph1.xlsx", "Graph 1"),
    #     ("Graph2.xlsx", "Graph 2"),
    #     # ("Compare1.xlsx", "Compare Graph 1"),
    #     # ("Compare2.xlsx", "Compare Graph 2"),
    # ]

    # ----------------------------------
    # Create merged workbook
    # ----------------------------------

    new_wb = Workbook()

    # Remove default sheet later after first copy
    first_sheet = True

    for file_path, sheet_name in files:

        wb = load_workbook(file_path)

        src_ws = wb.active

        if src_ws is None:
            print(f"Skipping {file_path}: No active sheet")
            continue

        if first_sheet:

            dst_ws = new_wb.active

            if dst_ws is None:
                raise ValueError("Failed to create worksheet")

            dst_ws.title = sheet_name

            first_sheet = False

        else:

            dst_ws = new_wb.create_sheet(title=sheet_name)

        copy_sheet(src_ws, dst_ws)

    # ----------------------------------
    # Save
    # ----------------------------------
    filename = "mergedExcel.xlsx"
    file_path = os.path.join(UPLOAD_FOLDER, filename)
    new_wb.save(file_path)

    print("Merged successfully!")

    return file_path